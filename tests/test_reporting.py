"""Tests for sake_rice_inspection.reporting."""

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from PIL import Image

from sake_rice_inspection.reporting import create_unified_prediction_excel


def make_image(path):
    Image.fromarray(np.zeros((10, 10, 3), dtype=np.uint8)).save(path)


def test_create_unified_prediction_excel_writes_one_row_per_result(tmp_path):
    img_dir = tmp_path / "crops"
    img_dir.mkdir()
    img1 = img_dir / "grain_0001.jpg"
    img2 = img_dir / "grain_0002.jpg"
    make_image(img1)
    make_image(img2)

    all_fold_results = [
        {"fold": 0, "img_path": str(img1), "true_labels": [1, 0], "true_label_str": "A", "probs": np.array([0.9, 0.1])},
        {"fold": 0, "img_path": str(img2), "true_labels": [0, 1], "true_label_str": "B", "probs": np.array([0.2, 0.8])},
    ]

    out_dir = tmp_path / "experiment"
    excel_path = create_unified_prediction_excel(all_fold_results, ["A", "B"], str(out_dir), image_root=str(img_dir))

    wb = load_workbook(excel_path)
    ws = wb.active

    assert ws.cell(row=1, column=1).value == "Image"
    assert ws.cell(row=2, column=2).value == "grain_0001.jpg"
    assert ws.cell(row=2, column=3).value == "A"
    assert ws.cell(row=2, column=4).value == "A"
    assert ws.cell(row=3, column=4).value == "B"
    assert ws.max_row == 3


def test_no_predicted_component_survives_excel_roundtrip_without_becoming_nan(tmp_path):
    # Regression test: writing the literal string "None" into a cell used to
    # silently become NaN the next time the workbook was read with
    # pandas.read_excel, because "None" is one of pandas' default na_values.
    img_dir = tmp_path / "crops"
    img_dir.mkdir()
    img1 = img_dir / "grain_0001.jpg"
    make_image(img1)

    all_fold_results = [
        {"fold": 0, "img_path": str(img1), "true_labels": [0, 0], "true_label_str": "no_prediction", "probs": np.array([0.1, 0.2])},
    ]

    out_dir = tmp_path / "experiment"
    excel_path = create_unified_prediction_excel(all_fold_results, ["A", "B"], str(out_dir), image_root=str(img_dir))

    df = pd.read_excel(excel_path)
    predicted_class = df["Predicted Class"].iloc[0]

    assert isinstance(predicted_class, str)
    assert not pd.isna(predicted_class)
