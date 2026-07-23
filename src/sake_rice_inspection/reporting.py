"""Excel report generation for multi-label prediction results."""

import os
from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image


def create_unified_prediction_excel(
    all_fold_results: List[Dict[str, Any]], class_names: List[str], out_dir: str, image_root: str = None,
) -> str:
    """Writes every fold's predictions, with thumbnails, to one Excel file.

    Args:
        all_fold_results (List[Dict[str, Any]]): Each entry needs keys
            `fold`, `img_path`, `true_labels` (multi-hot list), `true_label_str`,
            and `probs` (per-label probability array).
        class_names (List[str]): Name for each label column.
        out_dir (str): Directory to write `all_predictions.xlsx` into.
        image_root (str): If given, filenames are stored relative to this
            directory instead of as absolute paths.

    Returns:
        str: Path to the written Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "All_Predictions"

    headers = ["Image", "Filename", "True Class", "Predicted Class"] + [f"{cls}_score" for cls in class_names]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    for col_idx in range(5, 5 + len(class_names)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    row_idx = 2
    for result in all_fold_results:
        img_path = result["img_path"]
        probs = result["probs"]
        pred_labels = (probs >= 0.5).astype(int)
        # "None" is deliberately avoided: pandas' default na_values list treats the
        # literal string "None" (and "NA", "null", etc.) as missing data, so it would
        # silently turn into NaN the next time this workbook is read with read_excel.
        pred_label_str = "+".join(cls for i, cls in enumerate(class_names) if pred_labels[i] == 1) or "no_prediction"

        try:
            pil_img = Image.open(img_path).convert("RGB")
            pil_img.thumbnail((100, 100), Image.LANCZOS)

            buffer = BytesIO()
            pil_img.save(buffer, format="PNG")
            buffer.seek(0)

            xl_img = XLImage(buffer)
            xl_img.width = pil_img.width
            xl_img.height = pil_img.height
            ws.add_image(xl_img, f"A{row_idx}")
            ws.row_dimensions[row_idx].height = min(pil_img.height * 0.75, 75)
        except Exception:
            pass

        filename = os.path.relpath(img_path, image_root) if image_root else img_path
        ws.cell(row=row_idx, column=2, value=filename)
        ws.cell(row=row_idx, column=3, value=result["true_label_str"])
        ws.cell(row=row_idx, column=4, value=pred_label_str)
        for col_idx, prob in enumerate(probs, start=5):
            ws.cell(row=row_idx, column=col_idx, value=float(prob))

        row_idx += 1

    os.makedirs(out_dir, exist_ok=True)
    excel_path = os.path.join(out_dir, "all_predictions.xlsx")
    wb.save(excel_path)
    return excel_path
