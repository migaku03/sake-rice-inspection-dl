"""Sake Rice Quality Inspection System

This module provides an automated pipeline for classifying the visual quality of
sake rice grains using a hierarchical deep learning approach.

It utilizes Cellpose for individual grain segmentation and a cascade of 
pre-trained CNN models (EfficientNet, ResNet) to classify grains into 
18 complex phenotypic categories.
"""

import os
import time
import logging
from pathlib import Path
from io import BytesIO
from typing import Dict, List, Optional, Tuple, Any

import numpy as np
import torch
import torch.nn as nn
from PIL import Image, ImageOps
from torchvision import models, transforms
from skimage import io as skio
from skimage.measure import regionprops
from skimage.transform import rotate
from cellpose import models as cp_models
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)


class PadToSquare:
    """Pads a PIL Image to a square shape while maintaining the aspect ratio.

    Attributes:
        size (int): The target size (width and height) of the square image.
        padding_value (int): The pixel value used for padding (e.g., 0 for black).
    """

    def __init__(self, size: int = 224, padding_value: int = 0):
        self.size = size
        self.padding_value = padding_value
    
    def __call__(self, img: Image.Image) -> Image.Image:
        """Applies the padding to the image.

        Args:
            img (Image.Image): The input PIL Image.

        Returns:
            Image.Image: The padded square PIL Image.
        """
        w, h = img.size
        max_dim = max(w, h)
        
        # Scale down if the maximum dimension exceeds the target size
        if max_dim > self.size:
            scale = self.size / max_dim
            new_w = int(w * scale)
            new_h = int(h * scale)
            img = img.resize((new_w, new_h), Image.LANCZOS)
            w, h = new_w, new_h
        
        # Calculate padding to center the image
        left = (self.size - w) // 2
        top = (self.size - h) // 2
        right = self.size - w - left
        bottom = self.size - h - top
        
        img = ImageOps.expand(img, (left, top, right, bottom), fill=self.padding_value)
        return img


class SakeRiceSystem:
    """Hierarchical classification system for sake rice grains.

    This class manages the full pipeline: grain segmentation using Cellpose,
    preprocessing (rotation correction, background masking), and hierarchical
    classification using multiple deep learning models.

    Attributes:
        device (torch.device): The device (CPU or CUDA) used for computation.
        cp_model (cp_models.CellposeModel): The Cellpose model for segmentation.
        models (Dict[str, nn.Module]): Dictionary of pre-trained classification models.
        transform (transforms.Compose): Image transformation pipeline.
    """

    def __init__(self, model_paths: Dict[str, str], device: Optional[torch.device] = None):
        """Initializes the SakeRiceSystem with specified models.

        Args:
            model_paths (Dict[str, str]): Dictionary containing paths to the weights
                for each hierarchical level ('coarse', 'colored', 'opaque', 'partially_clouded').
            device (Optional[torch.device]): Target compute device. Defaults to CUDA if available.
        """
        self.device = device if device else torch.device("cuda" if torch.cuda.is_available() else "cpu")
        logger.info(f"Initializing system on device: {self.device}")
        
        # 1. Initialize Cellpose (diameter=50, model_type='cyto')
        self.cp_model = cp_models.CellposeModel(gpu=torch.cuda.is_available(), model_type='cyto')
        
        # 2. Define and load models for each hierarchical layer
        self.models = {
            'coarse': self._prepare_model(model_paths.get('coarse', ''), 7, 'efficientnet_b2'),
            'colored': self._prepare_model(model_paths.get('colored', ''), 2, 'resnet34'),
            'opaque': self._prepare_model(model_paths.get('opaque', ''), 2, 'resnet101'),
            'partially_clouded': self._prepare_model(model_paths.get('partially_clouded', ''), 4, 'resnet101') # Multi-label
        }
        
        # 3. Common Image Transformations
        self.transform = transforms.Compose([
            PadToSquare(size=224, padding_value=0),
            transforms.ToTensor(),
            transforms.Normalize([0.485, 0.456, 0.406], [0.229, 0.224, 0.225])
        ])
        
        # Class Definitions
        self.coarse_classes = [
            'Perfect Grain', 'Sprouted', 'Opaque White', 
            'Colored', 'Cracked', 'Partially Clouded', 'Green Immature'
        ]
        self.partial_components = ['Shinpaku', 'Base White', 'Back White', 'Belly White']
        self.allowed_partial_labels = [
            'Shinpaku', 'Back White', 'Belly White', 'Base White',
            'Shinpaku+Back White', 'Shinpaku+Belly White', 'Shinpaku+Base White',
            'Back White+Belly White', 'Base White+Belly White',
            'Shinpaku+Back White+Belly White'
        ]

    def _prepare_model(self, path: str, num_classes: int, arch: str) -> nn.Module:
        """Initializes and loads weights for a specific model architecture.

        Args:
            path (str): File path to the model weights (.pt file).
            num_classes (int): Number of output classes.
            arch (str): Architecture name (e.g., 'resnet34', 'efficientnet_b2').

        Returns:
            nn.Module: The loaded PyTorch model set to evaluation mode.
        """
        if 'resnet34' in arch:
            model = models.resnet34(weights=models.ResNet34_Weights.IMAGENET1K_V1)
            model.fc = nn.Linear(model.fc.in_features, num_classes)
        elif 'resnet101' in arch:
            model = models.resnet101(weights=models.ResNet101_Weights.IMAGENET1K_V1)
            model.fc = nn.Linear(model.fc.in_features, num_classes)
        elif 'efficientnet_b2' in arch:
            model = models.efficientnet_b2(weights=models.EfficientNet_B2_Weights.IMAGENET1K_V1)
            model.classifier[1] = nn.Linear(model.classifier[1].in_features, num_classes)
        else:
            raise ValueError(f"Unsupported architecture: {arch}")
        
        if not os.path.exists(path):
            logger.warning(f"Model file not found: {path}. Using random initialization.")
        else:
            try:
                state_dict = torch.load(path, map_location=self.device)
                model.load_state_dict(state_dict)
                logger.info(f"Successfully loaded model: {path}")
            except Exception as e:
                logger.warning(f"Failed to load model from {path}: {e}. Using pre-trained ImageNet weights.")
        
        model.to(self.device).eval()
        return model

    def apply_domain_logic_correction(self, probs: np.ndarray) -> str:
        """Applies domain knowledge correction to multi-label predictions.

        Resolves physically impossible combinations of rice traits by calculating
        penalized scores for allowed combinations.

        Args:
            probs (np.ndarray): Array of probabilities for each partial component.

        Returns:
            str: The corrected label string (e.g., 'Shinpaku+Back White').
        """
        prob_map = dict(zip(self.partial_components, probs))
        
        # Initial thresholding
        pred_comps = [c for c, p in prob_map.items() if p > 0.5]
        raw_label = "+".join(pred_comps) if pred_comps else "None"
        
        if raw_label in self.allowed_partial_labels:
            return raw_label
            
        # Apply scoring logic for unallowed combinations
        scores = {}
        for lbl in self.allowed_partial_labels:
            comps = lbl.split('+')
            vals = [prob_map[c] for c in comps]
            non_vals = [prob_map[c] for c in self.partial_components if c not in comps]
            
            # Score = Mean of components - Penalty (0.05 * Mean of non-components)
            score = np.mean(vals) - (0.05 * np.mean(non_vals) if non_vals else 0)
            scores[lbl] = score
            
        return max(scores, key=scores.get)

    def preprocess_grain(self, img: np.ndarray, prop: Any, mask: np.ndarray, margin: int = 20) -> Image.Image:
        """Extracts, rotates, and masks a single grain from the main image.

        Args:
            img (np.ndarray): The original full image.
            prop (Any): RegionProperties object from skimage for the specific grain.
            mask (np.ndarray): The full segmentation mask.
            margin (int): Margin around the bounding box.

        Returns:
            Image.Image: The preprocessed grain image.
        """
        label_id = prop.label
        
        # Bounding box with margin
        y1, x1, y2, x2 = prop.bbox
        y1 = max(y1 - margin, 0)
        x1 = max(x1 - margin, 0)
        y2 = min(y2 + margin, mask.shape[0])
        x2 = min(x2 + margin, mask.shape[1])
        
        img_crop = img[y1:y2, x1:x2]
        label_mask = (mask == label_id).astype(np.uint8)
        mask_crop = label_mask[y1:y2, x1:x2]
        
        # Rotation correction based on major axis orientation
        angle_deg = -np.rad2deg(prop.orientation)
        img_rot = rotate(img_crop, angle=angle_deg, resize=True, order=1, preserve_range=True)
        mask_rot = rotate(mask_crop, angle=angle_deg, resize=True, order=0, preserve_range=True)
        mask_rot = (mask_rot > 0.5).astype(np.uint8)
        
        # Trim based on rotated mask
        ys, xs = np.where(mask_rot > 0)
        if len(ys) == 0:
            out = np.clip(img_crop, 0, 255).astype(np.uint8)
            return Image.fromarray(out if out.ndim == 2 else np.uint8(out[:, :, :3]))
        
        min_y, max_y = ys.min(), ys.max() + 1
        min_x, max_x = xs.min(), xs.max() + 1
        img_trim = img_rot[min_y:max_y, min_x:max_x]
        mask_trim = mask_rot[min_y:max_y, min_x:max_x]
        
        # Apply mask to remove background
        if img_trim.ndim == 3:
            img_masked = img_trim * mask_trim[..., None]
        else:
            img_masked = img_trim * mask_trim
        
        if np.issubdtype(img.dtype, np.integer):
            out = np.clip(img_masked, 0, 255).astype(np.uint8)
        else:
            out = np.clip(img_masked * 255, 0, 255).astype(np.uint8)
        
        if out.ndim == 3:
            out = out[:, :, :3]
        
        return Image.fromarray(out)

    def _predict_single_grain(self, input_tensor: torch.Tensor, debug: bool = False, grain_id: Optional[int] = None) -> Tuple[str, float]:
        """Runs the hierarchical classification logic for a single grain.

        Args:
            input_tensor (torch.Tensor): The preprocessed image tensor.
            debug (bool): Whether to print debug information.
            grain_id (Optional[int]): ID of the grain for debug output.

        Returns:
            Tuple[str, float]: The final predicted label and its confidence score.
        """
        with torch.no_grad():
            # Layer 1: Coarse Classification
            out1 = self.models['coarse'](input_tensor)
            lv1_probs = torch.softmax(out1, dim=1).cpu().numpy()[0]
            lv1_idx = int(np.argmax(lv1_probs))
            lv1_label = self.coarse_classes[lv1_idx]
            lv1_score = float(lv1_probs[lv1_idx])

            if debug:
                logger.info(f"--- [Grain {grain_id}] Layer 1 (Coarse) ---")
                logger.info(f"Selected: {lv1_label} (score={lv1_score:.4f})")

            # Layer 2: Fine Classification based on Layer 1 result
            if lv1_label == 'Colored':
                out2 = self.models['colored'](input_tensor)
                out2_probs = torch.softmax(out2, dim=1).cpu().numpy()[0]
                sub_labels = ['Spotted', 'Diseased']
                sub_idx = int(np.argmax(out2_probs))
                final_label = sub_labels[sub_idx]
                final_score = float(lv1_score * out2_probs[sub_idx])
                
            elif lv1_label == 'Opaque White':
                out2 = self.models['opaque'](input_tensor)
                out2_probs = torch.softmax(out2, dim=1).cpu().numpy()[0]
                sub_labels = ['Milky White', 'Dead Grain']
                sub_idx = int(np.argmax(out2_probs))
                final_label = sub_labels[sub_idx]
                final_score = float(lv1_score * out2_probs[sub_idx])
                
            elif lv1_label == 'Partially Clouded':
                comp_probs = torch.sigmoid(self.models['partially_clouded'](input_tensor)).cpu().numpy()[0]
                final_label = self.apply_domain_logic_correction(comp_probs)

                if '+' in final_label:
                    parts = final_label.split('+')
                    comp_map = dict(zip(self.partial_components, comp_probs))
                    pw_score = float(np.mean([comp_map[p] for p in parts]))
                else:
                    pw_score = float(comp_probs[self.partial_components.index(final_label)])

                final_score = float(lv1_score * pw_score)
                
            else:
                final_label = lv1_label
                final_score = lv1_score

            if debug:
                logger.info(f"Final Prediction: {final_label} (score={final_score:.4f})")

        return final_label, final_score

    def predict_image_to_excel(self, img_path: str, excel_path: Optional[str] = None, area_threshold: int = 1500) -> Tuple[List[Dict[str, Any]], str]:
        """Processes an image, classifies all grains, and exports results to Excel.

        Args:
            img_path (str): Path to the input image.
            excel_path (Optional[str]): Path to save the Excel file. Defaults to outputs/predictions/.
            area_threshold (int): Minimum pixel area for a grain to be processed.

        Returns:
            Tuple[List[Dict[str, Any]], str]: A list of result dictionaries and the path to the Excel file.
        """
        img = skio.imread(img_path)
        masks, _, _ = self.cp_model.eval(img, channels=[0, 0], diameter=50)
        props = regionprops(masks, intensity_image=img)

        if excel_path is None:
            out_dir = Path("outputs/predictions")
            out_dir.mkdir(parents=True, exist_ok=True)
            excel_path = str(out_dir / f"{Path(img_path).stem}_predictions.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Final_Predictions"

        headers = ["Image", "Grain ID", "Predicted Class", "Final Score"]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 24
        ws.column_dimensions['D'].width = 14

        results = []
        row_idx = 2

        for idx, prop in enumerate(props):
            if prop.area < area_threshold:
                continue

            grain_img = self.preprocess_grain(img, prop, masks)
            input_tensor = self.transform(grain_img).unsqueeze(0).to(self.device)
            
            enable_debug = (idx < 5) # Debug output for the first 5 grains
            final_label, final_score = self._predict_single_grain(input_tensor, debug=enable_debug, grain_id=int(prop.label))

            results.append({
                "id": int(prop.label),
                "label": final_label,
                "score": float(final_score)
            })

            # Embed thumbnail into Excel
            try:
                thumb = grain_img.convert("RGB")
                thumb.thumbnail((100, 100), Image.LANCZOS)
                img_buffer = BytesIO()
                thumb.save(img_buffer, format='PNG')
                img_buffer.seek(0)

                xl_img = XLImage(img_buffer)
                xl_img.width = thumb.width
                xl_img.height = thumb.height
                ws.add_image(xl_img, f"A{row_idx}")
                ws.row_dimensions[row_idx].height = min(thumb.height * 0.75, 75)
            except Exception as exc:
                logger.error(f"Failed to embed image for ID={prop.label}: {exc}")

            ws.cell(row=row_idx, column=2, value=int(prop.label))
            ws.cell(row=row_idx, column=3, value=final_label)
            ws.cell(row=row_idx, column=4, value=float(final_score))
            row_idx += 1

        excel_parent = os.path.dirname(excel_path)
        if excel_parent:
            os.makedirs(excel_parent, exist_ok=True)
        wb.save(excel_path)

        return results, excel_path


if __name__ == "__main__":
    """Example execution script for the hierarchical classification pipeline.
    
    This block demonstrates how to initialize the system with model weights
    and process a sample image. Paths are configured relative to the repository root.
    """
    start_time = time.time()
    
    # Configure base paths
    base_dir = Path(__file__).resolve().parent.parent
    
    # Define relative paths for public repository structure
    input_image_path = str(base_dir / "data" / "sample" / "sample_grains.jpg")
    
    # Define model paths (Note: Actual weights are excluded from repo, use placeholders or instructions)
    model_weights = {
        'coarse': str(base_dir / 'models' / 'coarse_efficientnet_b2.pt'),
        'colored': str(base_dir / 'models' / 'colored_resnet34.pt'),
        'opaque': str(base_dir / 'models' / 'opaque_resnet101.pt'),
        'partially_clouded': str(base_dir / 'models' / 'partwhite_resnet101.pt')
    }
    
    logger.info("Initializing Sake Rice Quality Inspection System...")
    
    # Initialize system
    system = SakeRiceSystem(model_paths=model_weights)
    
    # Execute prediction
    if os.path.exists(input_image_path):
        logger.info(f"Processing image: {input_image_path}")
        all_grains, output_excel = system.predict_image_to_excel(input_image_path)

        elapsed_time = time.time() - start_time
        minutes, seconds = divmod(elapsed_time, 60)
        
        logger.info(f"Processing Complete: {len(all_grains)} grains analyzed in {int(minutes)}m {seconds:.1f}s")
        logger.info(f"Results saved to: {output_excel}")
    else:
        logger.warning(f"Sample image not found at {input_image_path}. Please add an image to test the pipeline.")
